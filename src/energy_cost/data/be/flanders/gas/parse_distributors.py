#!/usr/bin/env python3
"""
Parse Flemish gas distributor tariff Excel files and write YAML entries.

Source:
    https://www.vlaamsenutsregulator.be/elektriciteit-en-aardgas/nettarieven/hoeveel-bedragen-de-distributienettarieven

Usage:
    python parse_distributors.py <excel_path> [--year <year>] [--output-dir <dir>]

The year is inferred from the Excel filename when not specified.
For each distributor, a YAML file is created/updated in the 'distributors'
subdirectory next to the Excel file (or --output-dir if provided).

Existing entries for the same year are overwritten; other years are preserved.

Only the use case 'niet-telegemeten klanten met digitale gasmeter exclusief BTW'
is extracted: T1 (0–5 MWh/year), T2 (5–150 MWh/year), T3 (150+ MWh/year).

The all-in proportional cost per MWh is the sum of:
  - Basistarief proportionele term       (row 13, EUR/kWh)
  - ODV tarief                           (row 23, EUR/kWh)
  - Lasten niet-gekapitaliseerde pensioen (row 26, EUR/kWh)
  - Retributies/overige heffingen        (row 27, EUR/kWh)
converted to EUR/MWh.
"""

import argparse
import os
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl
import yaml

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

# Map sheet-name prefix → YAML filename stem (one file per distributor)
SHEET_TO_FILE: dict[str, str] = {
    "FA": "fluvius_antwerpen",
    "FHV": "fluvius_halle_vilvoorde",
    "FI": "fluvius_imewo",
    "FK": "fluvius_kempen",
    "FL": "fluvius_limburg",
    "FMV": "fluvius_midden_vlaanderen",
    "FW": "fluvius_west",
    "FZD": "fluvius_zenne_dijle",
}

# Band upper limits in MWh (matching kWh thresholds in the Excel / 1000)
BAND_T1_MWH = 5  # 0 – 5 000 kWh/year
BAND_T2_MWH = 150  # 5 001 – 150 000 kWh/year
BAND_T3_MWH = 1000  # 150 001 – 1 000 000 kWh/year
# T4 is open-ended (> 1 000 000 kWh/year)

# ---------------------------------------------------------------------------
# Layout
#
# The gas Excel format is consistent across all years inspected (2024–2026).
# Row indices are 0-based; column indices for T1/T2/T3 are 3/4/5.
# ---------------------------------------------------------------------------

ROW_VASTE = 12  # 'Vaste term'           EUR/jaar  – fixed cost per band
ROW_PROP = 13  # 'Proportionele term'   EUR/kWh   – network-use rate
ROW_DATABEHEER = 21  # 'Jaaropname'           EUR/jaar  – digital-meter mgmt fee
ROW_ODV = 23  # ODV tarief             EUR/kWh   – public-service obligation
ROW_HEFF1 = 26  # Lasten pensioenen      EUR/kWh   – pension levy
ROW_HEFF2 = 27  # Retributies/overige    EUR/kWh   – other levies

# Column index for T1, T2, T3, T4 (0-based)
COL_T1 = 3
COL_T2 = 4
COL_T3 = 5
COL_T4 = 6

CET = timezone(timedelta(hours=1))

# ---------------------------------------------------------------------------
# YAML serialisation
# ---------------------------------------------------------------------------


class _Dumper(yaml.Dumper):
    """PyYAML dumper that serialises datetime as ISO 8601 strings."""


def _represent_datetime(dumper: yaml.Dumper, data: datetime) -> yaml.ScalarNode:
    return dumper.represent_scalar("tag:yaml.org,2002:str", data.isoformat())


_Dumper.add_representer(datetime, _represent_datetime)

# ---------------------------------------------------------------------------
# Tariff extraction
# ---------------------------------------------------------------------------


def extract_tariffs(ws) -> dict:
    """
    Extract residential digital-meter tariff values from a GAS Afname worksheet.

    Returns a dict with:
      fixed_distribution_fee        : list[float]  – EUR/year for T1, T2, T3, T4
      proportional_distribution_fee : list[float]  – EUR/MWh for T1, T2, T3, T4
      public_service_obligation     : list[float]  – EUR/MWh for T1, T2, T3, T4
      pension_levy                  : list[float]  – EUR/MWh for T1, T2, T3, T4
      other_levies                  : list[float]  – EUR/MWh for T1, T2, T3, T4
      databeheer                    : float        – EUR/year (Jaaropname)
    """
    rows = list(ws.iter_rows(values_only=True))

    fixed_distribution_fee = []
    proportional_distribution_fee = []
    public_service_obligation = []
    pension_levy = []
    other_levies = []

    for col in (COL_T1, COL_T2, COL_T3, COL_T4):
        fixed_distribution_fee.append(rows[ROW_VASTE][col])
        # Convert EUR/kWh → EUR/MWh
        proportional_distribution_fee.append(round((rows[ROW_PROP][col] or 0.0) * 1000, 4))
        public_service_obligation.append(round((rows[ROW_ODV][col] or 0.0) * 1000, 4))
        pension_levy.append(round((rows[ROW_HEFF1][col] or 0.0) * 1000, 4))
        other_levies.append(round((rows[ROW_HEFF2][col] or 0.0) * 1000, 4))

    databeheer = rows[ROW_DATABEHEER][COL_T1]

    return {
        "fixed_distribution_fee": fixed_distribution_fee,
        "proportional_distribution_fee": proportional_distribution_fee,
        "public_service_obligation": public_service_obligation,
        "pension_levy": pension_levy,
        "other_levies": other_levies,
        "databeheer": databeheer,
    }


def _banded_mwh(values: list[float], yearly: bool = False) -> dict:
    """Build a banded consumption formula dict from a 4-element list [T1, T2, T3, T4]."""
    formula = (lambda v: {"period": "yearly", "constant_cost": v}) if yearly else (lambda v: {"constant_cost": v})
    return {
        "mode": "banded",
        "band_period": "P1Y",
        "bands": [
            {"up_to": BAND_T1_MWH, "formula": formula(values[0])},
            {"up_to": BAND_T2_MWH, "formula": formula(values[1])},
            {"up_to": BAND_T3_MWH, "formula": formula(values[2])},
            {"formula": formula(values[3])},
        ],
    }


def build_entry(year: int, tariffs: dict) -> dict:
    """Build the YAML-serialisable dict for one year's gas tariff entry."""
    return {
        "start": datetime(year, 1, 1, 0, 0, 0, tzinfo=CET),
        "periodic": {
            "data_management": {
                "period": "yearly",
                "constant_cost": tariffs["databeheer"],
            }
        },
        "consumption": {
            "fixed_distribution_fee": _banded_mwh(tariffs["fixed_distribution_fee"], yearly=True),
            "proportional_distribution_fee": _banded_mwh(tariffs["proportional_distribution_fee"]),
            "public_service_obligation": _banded_mwh(tariffs["public_service_obligation"]),
            "pension_levy": _banded_mwh(tariffs["pension_levy"]),
            "other_levies": _banded_mwh(tariffs["other_levies"]),
        },
    }


# ---------------------------------------------------------------------------
# YAML I/O helpers
# ---------------------------------------------------------------------------


def _entry_year(entry: dict) -> int:
    """Return the year from an entry's 'start' field (datetime or ISO string)."""
    start = entry.get("start")
    if isinstance(start, datetime):
        return start.year
    if isinstance(start, str):
        return int(start[:4])
    raise ValueError(f"Cannot determine year from start value: {start!r}")


def load_entries(path: Path) -> list:
    if not path.exists():
        return []
    with open(path) as f:
        data = yaml.safe_load(f)
    return data if data is not None else []


def save_entries(path: Path, entries: list) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as f:
        yaml.dump(
            entries,
            f,
            Dumper=_Dumper,
            default_flow_style=False,
            allow_unicode=True,
            sort_keys=False,
        )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def infer_year(excel_path: str) -> int:
    m = re.search(r"\b(20\d{2})\b", os.path.basename(excel_path))
    if not m:
        raise ValueError(f"Cannot infer year from filename: {excel_path!r}")
    return int(m.group(1))


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Parse Flemish gas distributor tariff Excel to YAML entries.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("excel", help="Path to the gas distributor tariff Excel file")
    parser.add_argument(
        "--year",
        type=int,
        default=None,
        help="Tariff year (inferred from the filename when omitted)",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Directory for YAML output files (defaults to <excel_dir>/distributors)",
    )
    args = parser.parse_args()

    year = args.year if args.year is not None else infer_year(args.excel)
    output_dir = Path(args.output_dir) if args.output_dir else Path(args.excel).parent / "distributors"

    print(f"Processing year {year} from: {args.excel}")
    print(f"Output directory: {output_dir}\n")

    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)
    processed = 0

    for sheet_name in wb.sheetnames:
        if not sheet_name.endswith("GAS Afname"):
            continue
        prefix = sheet_name.split()[0]
        if prefix not in SHEET_TO_FILE:
            print(f"  WARNING: unknown sheet prefix {prefix!r}, skipping {sheet_name!r}")
            continue

        ws = wb[sheet_name]
        tariffs = extract_tariffs(ws)
        new_entry = build_entry(year, tariffs)

        yaml_path = output_dir / f"{SHEET_TO_FILE[prefix]}.yml"
        entries = load_entries(yaml_path)
        # Remove existing entry for this year (overwrite semantics)
        entries = [e for e in entries if _entry_year(e) != year]
        entries.append(new_entry)
        entries.sort(key=_entry_year)
        save_entries(yaml_path, entries)
        print(f"  {SHEET_TO_FILE[prefix]}.yml  updated (year {year})")
        processed += 1

    print(f"\nDone. {processed} distributor(s) processed.")


if __name__ == "__main__":
    main()
