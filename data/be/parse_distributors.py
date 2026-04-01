#!/usr/bin/env python3
"""
Parse Flemish distributor tariff Excel files and write YAML entries.

Source:
    https://www.vlaamsenutsregulator.be/elektriciteit-en-aardgas/nettarieven/hoeveel-bedragen-de-distributienettarieven

Usage:
    python parse_distributors.py <excel_path> [--year <year>] [--output-dir <dir>]

The year is inferred from the Excel filename when not specified.
For each distributor, a YAML file is created/updated in the 'distributors'
subdirectory of the Excel file (or --output-dir if provided).

Existing entries for the same year are overwritten; other years are preserved.
Only the use case 'digitale meter exclusief BTW' (piekmeting on laagspanningsnet)
is extracted.
"""

import argparse
import os
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl
import yaml

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

# Map sheet-name prefix → YAML filename stem (one file per distributor)
SHEET_TO_FILE: dict[str, str] = {
    "FA": "fluvius_antwerpen",
    "FHV": "fluvius_halle_vilvoorde",
    "FI": "fluvius_imewo",
    "FK": "fluvius_kempen",
    "FL": "fluvius_limburg",
    "FMV": "fluvius_midden_vlaanderen",
    "FW": "fluvius_west",
    "FZD": "fluvius_zenne_dijle",
}

# Minimum capacity threshold for the lower price band (2.5 kW = 0.0025 MW)
MIN_CAPACITY_MW = 0.0025

# ---------------------------------------------------------------------------
# Layout detection
#
# The Excel format changed between 2024 and 2025:
#
#  2025/2026 format – 'Laagspanningsnet' at row 4, col 13 (0-based)
#    col 13 = LS piekmeting (digital meter)
#
#  2024 format – 'LS' at row 4, col 14 (0-based)
#    col 14 = LS piekmeting (digital meter)
#    Databeheer and ODV/toeslagen rows are also shifted relative to 2025+.
# ---------------------------------------------------------------------------

# Layout config: maps a format-key to (piekmeting_col, row_indices…)
# Row indices are 0-based.
_LAYOUTS = {
    # 2025 / 2026 – detected when rows[3][13] == 'Laagspanningsnet'
    "new": dict(
        col=13,
        row_cap=14,  # Row 15 – Gemiddelde maandpiek  EUR/kW/jaar
        row_kwh_net=16,  # Row 17 – kWh-tarief netgebruik  EUR/kWh
        row_databeheer=26,  # Row 27 – Databeheer laagspanningsnet  EUR/jaar
        row_odv_norm=29,  # Row 30 – ODV kWh-tarief normaal  EUR/kWh
        row_odv_nacht=30,  # Row 31 – ODV kWh-tarief exclusief nacht  EUR/kWh
        row_toeslagen=32,  # Row 33 – Toeslagen  EUR/kWh
    ),
    # 2024 – different column mapping and extra 'overige transmissienetkosten' row
    "old": dict(
        col=14,
        row_cap=14,  # Row 15 – Gemiddelde maandpiek  EUR/kW/jaar
        row_kwh_net=16,  # Row 17 – kWh-tarief netgebruik  EUR/kWh
        row_databeheer=27,  # Row 28 – Databeheer LS per kwartier  EUR/jaar
        row_odv_norm=31,  # Row 32 – ODV kWh-tarief normaal  EUR/kWh
        row_odv_nacht=32,  # Row 33 – ODV kWh-tarief exclusief nacht  EUR/kWh
        row_toeslagen=34,  # Row 35 – Toeslagen  EUR/kWh
    ),
}

CET = timezone(timedelta(hours=1))

# ---------------------------------------------------------------------------
# YAML serialisation
# ---------------------------------------------------------------------------


class _Dumper(yaml.Dumper):
    """PyYAML dumper that serialises datetime as ISO 8601 strings."""


def _represent_datetime(dumper: yaml.Dumper, data: datetime) -> yaml.ScalarNode:
    return dumper.represent_scalar("tag:yaml.org,2002:str", data.isoformat())


_Dumper.add_representer(datetime, _represent_datetime)

# ---------------------------------------------------------------------------
# Tariff extraction
# ---------------------------------------------------------------------------


def extract_tariffs(ws) -> dict:
    """
    Extract the 'digitale meter exclusief BTW' tariff values from an
    Afname worksheet and return them as a dict ready for build_entry().

    Values read from the LS piekmeting (digital meter) column:
      - capacity_per_month : EUR/MW/month  (converted from EUR/kW/year)
      - min_band_cost      : EUR/month for the ≤0.0025 MW band
      - consumption_all    : EUR/MWh  (netgebruik + ODV normaal + toeslagen)
      - consumption_night  : EUR/MWh  (netgebruik + ODV nacht  + toeslagen)
      - databeheer         : EUR/year
    """
    rows = list(ws.iter_rows(values_only=True))

    # Detect layout: 'Laagspanningsnet' at row 4 col 13 → 2025+ format
    layout = _LAYOUTS["new"] if rows[3][13] == "Laagspanningsnet" else _LAYOUTS["old"]
    col = layout["col"]

    cap_kw_year = rows[layout["row_cap"]][col]
    kwh_net = rows[layout["row_kwh_net"]][col]
    databeheer = rows[layout["row_databeheer"]][col]
    odv_normaal = rows[layout["row_odv_norm"]][col]
    odv_nacht = rows[layout["row_odv_nacht"]][col]
    toeslagen = rows[layout["row_toeslagen"]][col]

    # Convert capacity: EUR/kW/year → EUR/MW/month
    cap_per_month = cap_kw_year * 1000 / 12
    min_band = MIN_CAPACITY_MW * cap_per_month

    # Sum kWh components and convert EUR/kWh → EUR/MWh
    consumption_all = (kwh_net + odv_normaal + toeslagen) * 1000
    consumption_night = (kwh_net + odv_nacht + toeslagen) * 1000

    return {
        "capacity_per_month": round(cap_per_month, 7),
        "min_band_cost": round(min_band, 7),
        "consumption_all": round(consumption_all, 4),
        "consumption_night": round(consumption_night, 4),
        "databeheer": databeheer,
    }


def build_entry(year: int, tariffs: dict) -> dict:
    """Build the YAML-serialisable dict for one year's tariff entry."""
    return {
        "start": datetime(year, 1, 1, 0, 0, 0, tzinfo=CET),
        "capacity": {
            "measurement_period": "PT15M",
            "billing_period": "P1M",
            "window_periods": 12,
            "formula": {
                "bands": [
                    {
                        "up_to": MIN_CAPACITY_MW,
                        "formula": {
                            "period": "monthly",
                            "constant_cost": tariffs["min_band_cost"],
                        },
                    },
                    {
                        "formula": {
                            "constant_cost": tariffs["capacity_per_month"],
                        }
                    },
                ]
            },
        },
        "consumption": {
            "all": {"constant_cost": tariffs["consumption_all"]},
            "night_only": {"constant_cost": tariffs["consumption_night"]},
        },
        "periodic": {
            "data_mamangement": {
                "period": "yearly",
                "constant_cost": tariffs["databeheer"],
            }
        },
    }


# ---------------------------------------------------------------------------
# YAML I/O helpers
# ---------------------------------------------------------------------------


def _entry_year(entry: dict) -> int:
    """Return the year from an entry's 'start' field (datetime or ISO string)."""
    start = entry.get("start")
    if isinstance(start, datetime):
        return start.year
    if isinstance(start, str):
        return int(start[:4])
    raise ValueError(f"Cannot determine year from start value: {start!r}")


def load_entries(path: Path) -> list:
    if not path.exists():
        return []
    with open(path) as f:
        data = yaml.safe_load(f)
    return data if data is not None else []


def save_entries(path: Path, entries: list) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as f:
        yaml.dump(
            entries,
            f,
            Dumper=_Dumper,
            default_flow_style=False,
            allow_unicode=True,
            sort_keys=False,
        )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def infer_year(excel_path: str) -> int:
    m = re.search(r"\b(20\d{2})\b", os.path.basename(excel_path))
    if not m:
        raise ValueError(f"Cannot infer year from filename: {excel_path!r}")
    return int(m.group(1))


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Parse Flemish distributor tariff Excel to YAML entries.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("excel", help="Path to the distributor tariff Excel file")
    parser.add_argument(
        "--year",
        type=int,
        default=None,
        help="Tariff year (inferred from the filename when omitted)",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help=("Directory for YAML output files (defaults to <excel_dir>/distributors)"),
    )
    args = parser.parse_args()

    year = args.year if args.year is not None else infer_year(args.excel)
    output_dir = Path(args.output_dir) if args.output_dir else Path(args.excel).parent / "distributors"

    print(f"Processing year {year} from: {args.excel}")
    print(f"Output directory: {output_dir}\n")

    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)
    processed = 0

    for sheet_name in wb.sheetnames:
        if not sheet_name.endswith("ELEK Afname"):
            continue
        prefix = sheet_name.split()[0]
        if prefix not in SHEET_TO_FILE:
            print(f"  WARNING: unknown sheet prefix {prefix!r}, skipping {sheet_name!r}")
            continue

        ws = wb[sheet_name]
        tariffs = extract_tariffs(ws)
        new_entry = build_entry(year, tariffs)

        yaml_path = output_dir / f"{SHEET_TO_FILE[prefix]}.yml"
        entries = load_entries(yaml_path)
        # Remove existing entry for this year (overwrite semantics)
        entries = [e for e in entries if _entry_year(e) != year]
        entries.append(new_entry)
        entries.sort(key=_entry_year)
        save_entries(yaml_path, entries)

        print(f"  {yaml_path.name:<40} ← {sheet_name}")
        processed += 1

    wb.close()
    print(f"\nDone. {processed} distributor files updated for year {year}.")


if __name__ == "__main__":
    main()
